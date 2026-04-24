"""
NRL Team Handicap Matrix Builder
Generates one sheet per NRL team showing handicap cover rate vs market implied 50%.
Source: /Users/elliotbladen/Downloads/nrl (5).xlsx (seasons 2022-2025)
Output: outputs/nrl_handicap_matrix.xlsx
        outputs/nrl_handicap_matrix.csv

COLUMNS (same structure as H2H and totals matrices)
-------
  Cover Rate %          — % of games the team covered the spread
  Market Implied %      — always 50.0 (the line IS the market's 50/50 point)
  Difference (pp)       — Cover Rate minus 50%, signed
  Edge % & Direction    — relative % deviation from implied, labelled covers/fades
  N (Games)             — sample size

A game is covered when: actual_margin (team perspective) + team_line > 0
Edge flag threshold: 15% relative deviation from the 50% implied baseline,
i.e. cover rate >= 57.5% or <= 42.5%.
"""

import csv
from datetime import datetime, timedelta, date
from collections import defaultdict

import ephem
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SOURCE_PATH = "/Users/elliotbladen/Downloads/nrl (5).xlsx"
OUTPUT_PATH = "/Users/elliotbladen/Betting_model/outputs/nrl_handicap_matrix.xlsx"
CSV_PATH    = "/Users/elliotbladen/Betting_model/outputs/nrl_handicap_matrix.csv"
SEASONS     = (2022, 2023, 2024, 2025)
MIN_SAMPLE  = 3
EDGE_FLAG_PCT    = 15.0
MOON_WINDOW_DAYS = 1


# ─────────────────────────────────────────────
#  Moon phase helpers
# ─────────────────────────────────────────────

def build_moon_sets(start_date: date, end_date: date):
    new_moons, full_moons = set(), set()
    d = ephem.Date(start_date - timedelta(days=30))
    end_ephem = ephem.Date(end_date + timedelta(days=30))
    while d < end_ephem:
        nm = ephem.next_new_moon(d)
        fm = ephem.next_full_moon(d)
        nm_date = ephem.Date(nm).datetime().date()
        fm_date = ephem.Date(fm).datetime().date()
        for delta in range(-MOON_WINDOW_DAYS, MOON_WINDOW_DAYS + 1):
            new_moons.add(nm_date + timedelta(days=delta))
            full_moons.add(fm_date + timedelta(days=delta))
        d = nm + 1
    return new_moons, full_moons


# ─────────────────────────────────────────────
#  Team name map
# ─────────────────────────────────────────────

TEAM_NAME_MAP = {
    "Brisbane Broncos":        "Brisbane Broncos",
    "Canberra Raiders":        "Canberra Raiders",
    "Canterbury Bulldogs":     "Canterbury Bulldogs",
    "Cronulla Sharks":         "Cronulla Sharks",
    "Dolphins":                "Dolphins",
    "Gold Coast Titans":       "Gold Coast Titans",
    "Manly Sea Eagles":        "Manly Sea Eagles",
    "Melbourne Storm":         "Melbourne Storm",
    "New Zealand Warriors":    "New Zealand Warriors",
    "Newcastle Knights":       "Newcastle Knights",
    "North QLD Cowboys":       "North QLD Cowboys",
    "Parramatta Eels":         "Parramatta Eels",
    "Penrith Panthers":        "Penrith Panthers",
    "South Sydney Rabbitohs":  "South Sydney Rabbitohs",
    "St George Dragons":       "St George Dragons",
    "Sydney Roosters":         "Sydney Roosters",
    "Wests Tigers":            "Wests Tigers",
}


# ─────────────────────────────────────────────
#  Data loading
# ─────────────────────────────────────────────

def load_data():
    wb = openpyxl.load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for raw in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        game_date_raw   = raw[0]
        kickoff_raw     = raw[1]
        home_team_raw   = raw[2]
        away_team_raw   = raw[3]
        venue           = raw[4]
        home_score      = raw[5]
        away_score      = raw[6]
        is_playoff      = bool(raw[7])
        # Prefer closing line; fall back to opening line
        hcap_home_close = raw[24] if raw[24] is not None else raw[21]

        if not game_date_raw or not hasattr(game_date_raw, "year"):
            continue
        if game_date_raw.year not in SEASONS:
            continue
        if home_score is None or away_score is None:
            continue
        if hcap_home_close is None:
            continue

        game_date = game_date_raw.date() if hasattr(game_date_raw, "date") else game_date_raw
        if kickoff_raw and hasattr(kickoff_raw, "hour") and not hasattr(kickoff_raw, "date"):
            dt = datetime.combine(game_date, kickoff_raw)
        elif kickoff_raw and hasattr(kickoff_raw, "date"):
            dt = datetime.combine(game_date, kickoff_raw.time())
        else:
            dt = datetime(game_date.year, game_date.month, game_date.day, 19, 0)

        home_team = TEAM_NAME_MAP.get(home_team_raw, home_team_raw)
        away_team = TEAM_NAME_MAP.get(away_team_raw, away_team_raw)

        rows.append({
            "dt":                 dt,
            "game_date":          game_date,
            "season":             game_date_raw.year,
            "home_team":          home_team,
            "away_team":          away_team,
            "venue":              venue or "Unknown",
            "home_score":         int(home_score),
            "away_score":         int(away_score),
            "is_playoff":         is_playoff,
            "actual_margin_home": int(home_score) - int(away_score),
            "hcap_home":          float(hcap_home_close),
        })

    wb.close()
    rows.sort(key=lambda r: r["dt"])
    return rows


def enrich_rows(rows):
    dates = [r["game_date"] for r in rows]
    new_moon_dates, full_moon_dates = build_moon_sets(min(dates), max(dates))

    for r in rows:
        dt = r["dt"]
        wd = dt.weekday()
        r["is_night"]    = dt.hour >= 18
        r["is_day"]      = dt.hour < 18
        r["is_thu_fri"]  = wd in (3, 4)
        r["is_saturday"] = wd == 5
        r["is_sunday"]   = wd == 6
        r["is_new_moon"] = r["game_date"] in new_moon_dates
        r["is_full_moon"]= r["game_date"] in full_moon_dates

    team_games_map: dict = defaultdict(list)
    for r in rows:
        team_games_map[r["home_team"]].append(r)
        team_games_map[r["away_team"]].append(r)

    for team, games in team_games_map.items():
        games.sort(key=lambda x: x["dt"])
        for i, g in enumerate(games):
            key_rest = f"rest__{team}"
            key_prev = f"prev_win__{team}"
            if i == 0:
                g[key_rest] = None
                g[key_prev] = None
            else:
                prev = games[i - 1]
                g[key_rest] = (g["game_date"] - prev["game_date"]).days
                if prev["home_team"] == team:
                    g[key_prev] = prev["home_score"] > prev["away_score"]
                else:
                    g[key_prev] = prev["away_score"] > prev["home_score"]

    return rows


# ─────────────────────────────────────────────
#  Stats computation
# ─────────────────────────────────────────────

def team_margin(game, team):
    if game["home_team"] == team:
        return game["actual_margin_home"]
    return -game["actual_margin_home"]


def team_line(game, team):
    """
    Handicap line from team's perspective.
    Negative = team is favoured. Positive = team is the underdog.
    """
    if game["home_team"] == team:
        return game["hcap_home"]
    return -game["hcap_home"]


def compute_stats(games, team):
    """
    Returns (cover_rate_pct, market_implied_pct, n) or None if below MIN_SAMPLE.
    market_implied_pct is always 50.0 — the handicap line IS the 50/50 point.
    """
    n = len(games)
    if n < MIN_SAMPLE:
        return None
    covers = sum(1 for g in games if team_margin(g, team) + team_line(g, team) > 0)
    cover_rate = round(covers / n * 100, 1)
    return cover_rate, 50.0, n


def edge_label(cover_rate, implied_pct):
    """Mirrors H2H edge_label exactly — relative % deviation from implied."""
    diff = cover_rate - implied_pct
    if implied_pct == 0:
        return round(diff, 1), 0.0, "", False
    edge_pct  = abs(diff / implied_pct) * 100
    direction = "covers" if diff > 0 else "fades"
    flag      = edge_pct >= EDGE_FLAG_PCT
    return round(diff, 1), round(edge_pct, 1), direction, flag


def team_games(all_rows, team):
    return [r for r in all_rows if r["home_team"] == team or r["away_team"] == team]


# ─────────────────────────────────────────────
#  Excel formatting  (identical to H2H / totals)
# ─────────────────────────────────────────────

HEADER_FILL      = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT      = Font(color="FFFFFF", bold=True, size=10)
SECTION_FILL     = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT     = Font(color="FFFFFF", bold=True, size=9)
LABEL_FILL       = PatternFill("solid", fgColor="D6E4F0")
ALT_ROW_FILL     = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL       = PatternFill("solid", fgColor="FFFFFF")
BLANK_FILL       = PatternFill("solid", fgColor="F2F2F2")
FLAG_FILL        = PatternFill("solid", fgColor="6CE58D")
STRONG_FLAG_FILL = PatternFill("solid", fgColor="00FF00")
THIN             = Side(style="thin", color="CCCCCC")
THIN_BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_HEADERS = [
    "Category",
    "Cover Rate %",
    "Market Implied %",
    "Difference (pp)",
    "Edge % & Direction",
    "N (Games)",
]


def style_header_row(ws, row):
    for c in range(1, 7):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_section_row(ws, row, label):
    ws.cell(row=row, column=1).value = label
    for c in range(1, 7):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER


def write_data_row(ws, row, label, stats, alt_row=False):
    bg = ALT_ROW_FILL if alt_row else WHITE_FILL

    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.fill = LABEL_FILL
    label_cell.font = Font(size=9)
    label_cell.alignment = Alignment(vertical="center", indent=1)
    label_cell.border = THIN_BORDER

    if stats is None:
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c, value="—")
            cell.fill = BLANK_FILL
            cell.font = Font(color="AAAAAA", size=9)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        return

    cover_rate, implied_pct, n = stats
    diff, edge_pct, direction, flag = edge_label(cover_rate, implied_pct)
    edge_text = f"{edge_pct}% {direction}" if edge_pct >= 1.0 else "—"

    for c_idx, val in enumerate([cover_rate, implied_pct, diff, edge_text, n], start=2):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = Font(bold=(c_idx == 5 and flag), size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if c_idx in (2, 3, 4):
            cell.fill = bg
            cell.number_format = "0.0"
        elif c_idx == 5:
            cell.fill = STRONG_FLAG_FILL if (flag and edge_pct >= 30) else FLAG_FILL if flag else bg
        else:
            cell.fill = bg


# ─────────────────────────────────────────────
#  Sheet builder
# ─────────────────────────────────────────────

def build_team_sheet(wb, team, all_rows, all_teams, csv_rows):
    ws = wb.create_sheet(title=team[:31])

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 12

    ws.merge_cells("A1:F1")
    tc = ws.cell(row=1, column=1, value=f"{team} — NRL Handicap Matrix (2022–2025)")
    tc.fill = PatternFill("solid", fgColor="0D2137")
    tc.font = Font(color="FFFFFF", bold=True, size=12)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(COL_HEADERS, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2)
    ws.row_dimensions[2].height = 30

    cur_row = 3
    g = team_games(all_rows, team)
    current_section = [""]

    def section(label):
        nonlocal cur_row
        current_section[0] = label
        style_section_row(ws, cur_row, label)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

    def row(label, games, alt=False):
        nonlocal cur_row
        stats = compute_stats(games, team)
        write_data_row(ws, cur_row, label, stats, alt_row=alt)
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1
        # Collect for CSV
        if stats is not None:
            cover_rate, implied_pct, n = stats
            diff, edge_pct, direction, flag = edge_label(cover_rate, implied_pct)
            csv_rows.append({
                "team": team, "section": current_section[0], "category": label,
                "cover_rate_pct": cover_rate, "market_implied_pct": implied_pct,
                "diff_pp": diff, "edge_pct": edge_pct, "direction": direction,
                "n": n, "flag": flag,
            })
        else:
            csv_rows.append({
                "team": team, "section": current_section[0], "category": label,
                "cover_rate_pct": "", "market_implied_pct": 50.0,
                "diff_pp": "", "edge_pct": "", "direction": "",
                "n": "", "flag": "",
            })

    # ── OVERALL ──────────────────────────────────────────────────
    section("OVERALL")
    row("Cover Rate — All Games", g)
    row("Cover Rate — Home",      [x for x in g if x["home_team"] == team], alt=True)
    row("Cover Rate — Away",      [x for x in g if x["away_team"] == team])

    # ── FAVOURITE vs UNDERDOG ────────────────────────────────────
    section("FAVOURITE vs UNDERDOG")
    row("As Favourite (line < 0)",    [x for x in g if team_line(x, team) < 0])
    row("As Underdog  (line > 0)",    [x for x in g if team_line(x, team) > 0], alt=True)
    row("Heavy Fav (line ≤ -9.5)",   [x for x in g if team_line(x, team) <= -9.5])
    row("Slight Fav (line -1 to -9)", [x for x in g if -9.0 <= team_line(x, team) <= -1.0], alt=True)
    row("Slight Dog (line +1 to +9)", [x for x in g if 1.0 <= team_line(x, team) <= 9.0])
    row("Big Dog    (line ≥ +9.5)",  [x for x in g if team_line(x, team) >= 9.5], alt=True)

    # ── TIME OF DAY ──────────────────────────────────────────────
    section("TIME OF DAY")
    row("Night Games (kick-off ≥ 18:00)", [x for x in g if x["is_night"]])
    row("Day Games (kick-off < 18:00)",   [x for x in g if x["is_day"]], alt=True)

    # ── DAY OF WEEK ──────────────────────────────────────────────
    section("DAY OF WEEK")
    row("Thursday / Friday Games", [x for x in g if x["is_thu_fri"]])
    row("Saturday Games",          [x for x in g if x["is_saturday"]], alt=True)
    row("Sunday Games",            [x for x in g if x["is_sunday"]])

    # ── FORM ─────────────────────────────────────────────────────
    section("FORM")
    key_prev = f"prev_win__{team}"
    row("After a Win",  [x for x in g if x.get(key_prev) is True])
    row("After a Loss", [x for x in g if x.get(key_prev) is False], alt=True)

    # ── REST ─────────────────────────────────────────────────────
    section("REST")
    key_rest = f"rest__{team}"
    row("Short Rest (≤ 6 days)",  [x for x in g if x.get(key_rest) is not None and x[key_rest] <= 6])
    row("Normal Rest (7-9 days)", [x for x in g if x.get(key_rest) is not None and 7 <= x[key_rest] <= 9], alt=True)
    row("Long Rest (≥ 10 days)",  [x for x in g if x.get(key_rest) is not None and x[key_rest] >= 10])

    # ── MOON PHASE ───────────────────────────────────────────────
    section("MOON PHASE")
    row("New Moon (±1 day)",  [x for x in g if x["is_new_moon"]])
    row("Full Moon (±1 day)", [x for x in g if x["is_full_moon"]], alt=True)

    # ── BY MONTH ─────────────────────────────────────────────────
    section("BY MONTH")
    nrl_months = [
        (3, "March"), (4, "April"),     (5, "May"),       (6, "June"),
        (7, "July"),  (8, "August"),    (9, "September"), (10, "October"),
    ]
    for i, (m, mname) in enumerate(nrl_months):
        row(mname, [x for x in g if x["dt"].month == m], alt=(i % 2 == 1))

    # ── HEAD TO HEAD vs OPPONENT ─────────────────────────────────
    section("HEAD TO HEAD vs OPPONENT")
    opponents = sorted(t for t in all_teams if t != team)
    for i, opp in enumerate(opponents):
        opp_games = [x for x in g if x["home_team"] == opp or x["away_team"] == opp]
        row(f"vs {opp}", opp_games, alt=(i % 2 == 1))

    # ── BY VENUE ─────────────────────────────────────────────────
    section("BY VENUE")
    venues_played = sorted(set(x["venue"] for x in g))
    for i, venue in enumerate(venues_played):
        row(venue, [x for x in g if x["venue"] == venue], alt=(i % 2 == 1))

    ws.freeze_panes = "B3"


# ─────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────

def main():
    print("Loading data...")
    all_rows = load_data()
    print(f"  Loaded {len(all_rows)} games (seasons {SEASONS})")

    season_counts = defaultdict(int)
    for r in all_rows:
        season_counts[r["season"]] += 1
    for s in sorted(season_counts):
        print(f"    {s}: {season_counts[s]} games")

    print("Enriching rows (moon, rest, form)...")
    all_rows = enrich_rows(all_rows)

    all_teams = sorted(
        set(r["home_team"] for r in all_rows) | set(r["away_team"] for r in all_rows)
    )
    print(f"  {len(all_teams)} teams")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    csv_rows = []

    print("Building sheets...")
    for team in all_teams:
        print(f"  {team}...")
        build_team_sheet(wb, team, all_rows, all_teams, csv_rows)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")

    csv_fields = [
        "team", "section", "category",
        "cover_rate_pct", "market_implied_pct", "diff_pp",
        "edge_pct", "direction", "n", "flag",
    ]
    with open(CSV_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fields)
        writer.writeheader()
        writer.writerows(csv_rows)
    print(f"Saved: {CSV_PATH}")
    print(f"\n{len(csv_rows)} rows written.")


if __name__ == "__main__":
    main()
