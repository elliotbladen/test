"""
EPL Over/Under 2.5 Goals Matrix Builder
One sheet per EPL team — over rate vs market implied probability.
Source: four CSVs for 2021-22, 2022-23, 2023-24, 2024-25
Output: outputs/epl_ou25_matrix.xlsx

Metric columns:
  Actual Over Rate %  — % of games that ended over 2.5 goals
  Market Implied Over % — 1 / Pinnacle closing over odds (fallback: Bet365 closing)
  Difference (pp)      — actual minus implied, in percentage points
  Edge % & Direction   — (diff / implied) * 100, flagged green if >= 15%
  N (Games)
"""

import csv
from datetime import datetime, timedelta, date
from collections import defaultdict

import ephem
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SOURCE_FILES = {
    "2021-22": "/Users/elliotbladen/Downloads/E0_2021_22_with_managers_FINAL.csv",
    "2022-23": "/Users/elliotbladen/Downloads/E0 (7) (1).csv",
    "2023-24": "/Users/elliotbladen/Downloads/E0_2023_24_with_managers_FIXED.csv",
    "2024-25": "/Users/elliotbladen/Downloads/E0_2024_25_with_managers_CURRENT.csv",
}
OUTPUT_PATH = "/Users/elliotbladen/Betting_model/outputs/epl_ou25_matrix.xlsx"
MIN_SAMPLE = 3
EDGE_FLAG_PCT = 15.0
MOON_WINDOW_DAYS = 1


# ─────────────────────────────────────────────
#  Moon helpers
# ─────────────────────────────────────────────

def build_moon_sets(start_date: date, end_date: date):
    new_moons, full_moons = set(), set()
    d = ephem.Date(start_date - timedelta(days=30))
    end_ephem = ephem.Date(end_date + timedelta(days=30))
    while d < end_ephem:
        nm = ephem.next_new_moon(d)
        fm = ephem.next_full_moon(d)
        nm_d = ephem.Date(nm).datetime().date()
        fm_d = ephem.Date(fm).datetime().date()
        for delta in range(-MOON_WINDOW_DAYS, MOON_WINDOW_DAYS + 1):
            new_moons.add(nm_d + timedelta(days=delta))
            full_moons.add(fm_d + timedelta(days=delta))
        d = nm + 1
    return new_moons, full_moons


# ─────────────────────────────────────────────
#  Data loading
# ─────────────────────────────────────────────

def safe_float(val):
    """For odds — returns None if zero or unparseable."""
    try:
        f = float(val)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def safe_score(val):
    """For goals — 0 is a valid score."""
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def load_data():
    rows = []
    for season_label, filepath in SOURCE_FILES.items():
        with open(filepath, encoding="utf-8", errors="replace") as fh:
            reader = csv.DictReader(fh)
            for raw in reader:
                date_str = raw.get("Date", "").strip()
                time_str = raw.get("Time", "").strip()
                home_team = raw.get("HomeTeam", "").strip()
                away_team = raw.get("AwayTeam", "").strip()

                if not date_str or not home_team or not away_team:
                    continue

                try:
                    game_date = datetime.strptime(date_str, "%d/%m/%Y").date()
                except ValueError:
                    continue

                # Parse kickoff time
                try:
                    t = datetime.strptime(time_str, "%H:%M").time() if time_str else None
                except ValueError:
                    t = None
                dt = datetime.combine(game_date, t) if t else datetime(
                    game_date.year, game_date.month, game_date.day, 15, 0
                )

                fthg = safe_score(raw.get("FTHG"))
                ftag = safe_score(raw.get("FTAG"))
                if fthg is None or ftag is None:
                    continue

                total_goals = int(fthg) + int(ftag)
                over_25 = total_goals > 2  # strictly over 2.5

                # FTR for form tracking
                ftr = raw.get("FTR", "").strip()  # H, D, A

                # Market: prefer Pinnacle closing, fallback Bet365 closing
                pc_over = safe_float(raw.get("PC>2.5"))
                b365c_over = safe_float(raw.get("B365C>2.5"))
                market_over_odds = pc_over if pc_over else b365c_over

                if market_over_odds is None:
                    continue

                market_implied_over = 1.0 / market_over_odds  # raw implied prob (includes margin)

                rows.append({
                    "season": season_label,
                    "dt": dt,
                    "game_date": game_date,
                    "home_team": home_team,
                    "away_team": away_team,
                    "total_goals": total_goals,
                    "over_25": over_25,
                    "ftr": ftr,
                    "market_implied_over": market_implied_over,
                    "market_over_odds": market_over_odds,
                })

    rows.sort(key=lambda r: r["dt"])
    return rows


def enrich_rows(rows):
    dates = [r["game_date"] for r in rows]
    new_moon_dates, full_moon_dates = build_moon_sets(min(dates), max(dates))

    for r in rows:
        dt = r["dt"]
        wd = dt.weekday()  # 0=Mon, 1=Tue, 2=Wed, 3=Thu, 4=Fri, 5=Sat, 6=Sun
        r["is_night"] = dt.hour >= 18
        r["is_day"] = dt.hour < 18
        r["weekday_name"] = dt.strftime("%A")
        r["is_monday"] = wd == 0
        r["is_tue_wed"] = wd in (1, 2)
        r["is_thursday"] = wd == 3
        r["is_friday"] = wd == 4
        r["is_saturday"] = wd == 5
        r["is_sunday"] = wd == 6
        r["month"] = dt.month
        r["is_new_moon"] = r["game_date"] in new_moon_dates
        r["is_full_moon"] = r["game_date"] in full_moon_dates

    # Per-team rest days and prev result
    team_games_map: dict[str, list] = defaultdict(list)
    for r in rows:
        team_games_map[r["home_team"]].append(r)
        team_games_map[r["away_team"]].append(r)

    for team, games in team_games_map.items():
        games.sort(key=lambda x: x["dt"])
        for i, g in enumerate(games):
            k_rest = f"rest__{team}"
            k_prev = f"prev_result__{team}"
            if i == 0:
                g[k_rest] = None
                g[k_prev] = None
            else:
                prev = games[i - 1]
                g[k_rest] = (g["game_date"] - prev["game_date"]).days
                # Determine prev result for this team
                if prev["home_team"] == team:
                    g[k_prev] = prev["ftr"]  # H=win, D=draw, A=loss
                else:
                    inverse = {"H": "A", "A": "H", "D": "D"}
                    g[k_prev] = inverse.get(prev["ftr"], None)

    return rows


# ─────────────────────────────────────────────
#  Stats computation
# ─────────────────────────────────────────────

def compute_stats(games):
    """Returns (actual_over_pct, market_implied_over_pct, diff_pp, n) or None."""
    n = len(games)
    if n < MIN_SAMPLE:
        return None
    actual_rate = sum(1 for g in games if g["over_25"]) / n * 100
    avg_implied = sum(g["market_implied_over"] for g in games) / n * 100
    diff = actual_rate - avg_implied
    return round(actual_rate, 1), round(avg_implied, 1), round(diff, 1), n


def edge_label(actual_pct, implied_pct, diff):
    if implied_pct == 0:
        return 0.0, "", False
    edge_pct = abs(diff / implied_pct) * 100
    direction = "overs" if diff > 0 else "unders"
    flag = edge_pct >= EDGE_FLAG_PCT
    return round(edge_pct, 1), direction, flag


def team_games(all_rows, team):
    return [r for r in all_rows if r["home_team"] == team or r["away_team"] == team]


# ─────────────────────────────────────────────
#  Excel formatting
# ─────────────────────────────────────────────

HEADER_FILL      = PatternFill("solid", fgColor="1F3864")
HEADER_FONT      = Font(color="FFFFFF", bold=True, size=10)
SECTION_FILL     = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT     = Font(color="FFFFFF", bold=True, size=9)
LABEL_FILL       = PatternFill("solid", fgColor="D6E4F0")
ALT_ROW_FILL     = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL       = PatternFill("solid", fgColor="FFFFFF")
BLANK_FILL       = PatternFill("solid", fgColor="F2F2F2")
FLAG_FILL        = PatternFill("solid", fgColor="6CE58D")
STRONG_FLAG_FILL = PatternFill("solid", fgColor="00FF00")
THIN             = Side(style="thin", color="CCCCCC")
THIN_BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_HEADERS = [
    "Category",
    "Actual Over 2.5 %",
    "Market Implied Over %",
    "Difference (pp)",
    "Edge % & Direction",
    "N (Games)",
]

MONTH_NAMES = {
    8: "August", 9: "September", 10: "October", 11: "November", 12: "December",
    1: "January", 2: "February", 3: "March", 4: "April", 5: "May",
}


def style_header_row(ws, row):
    for c in range(1, 7):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_section_row(ws, row, label):
    ws.cell(row=row, column=1).value = label
    for c in range(1, 7):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER


def write_data_row(ws, row_idx, label, stats, alt_row=False):
    bg = ALT_ROW_FILL if alt_row else WHITE_FILL

    label_cell = ws.cell(row=row_idx, column=1, value=label)
    label_cell.fill = LABEL_FILL
    label_cell.font = Font(size=9)
    label_cell.alignment = Alignment(vertical="center", indent=1)
    label_cell.border = THIN_BORDER

    if stats is None:
        for c in range(2, 7):
            cell = ws.cell(row=row_idx, column=c, value="—")
            cell.fill = BLANK_FILL
            cell.font = Font(color="AAAAAA", size=9)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        return

    actual_pct, implied_pct, diff, n = stats
    edge_pct, direction, flag = edge_label(actual_pct, implied_pct, diff)
    edge_text = f"{edge_pct}% {direction} edge" if edge_pct >= 1.0 else "—"

    for c_idx, val in enumerate([actual_pct, implied_pct, diff, edge_text, n], start=2):
        cell = ws.cell(row=row_idx, column=c_idx, value=val)
        cell.font = Font(bold=(c_idx == 5 and flag), size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if c_idx in (2, 3, 4):
            cell.fill = bg
            cell.number_format = "0.0"
        elif c_idx == 5:
            if flag:
                cell.fill = STRONG_FLAG_FILL if edge_pct >= 30 else FLAG_FILL
            else:
                cell.fill = bg
        else:
            cell.fill = bg


# ─────────────────────────────────────────────
#  Sheet builder
# ─────────────────────────────────────────────

def build_team_sheet(wb, team, all_rows, all_teams):
    ws = wb.create_sheet(title=team[:31])

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 12

    # Title
    ws.merge_cells("A1:F1")
    seasons_in = sorted(set(r["season"] for r in all_rows
                            if r["home_team"] == team or r["away_team"] == team))
    title_text = f"{team} — EPL Over/Under 2.5 Goals Matrix ({', '.join(seasons_in)})"
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.fill = PatternFill("solid", fgColor="0D1F3C")
    tc.font = Font(color="FFFFFF", bold=True, size=12)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(COL_HEADERS, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2)
    ws.row_dimensions[2].height = 30

    cur_row = 3
    g = team_games(all_rows, team)

    def section(label):
        nonlocal cur_row
        style_section_row(ws, cur_row, label)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

    def row(label, games, alt=False):
        nonlocal cur_row
        write_data_row(ws, cur_row, label, compute_stats(games), alt_row=alt)
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

    # ── OVERALL ─────────────────────────────────────────────────
    section("OVERALL")
    row("All Games", g)
    row("Home Games", [x for x in g if x["home_team"] == team], alt=True)
    row("Away Games", [x for x in g if x["away_team"] == team])

    # ── TIME OF DAY ──────────────────────────────────────────────
    section("TIME OF DAY")
    row("Night Games (kick-off ≥ 18:00)", [x for x in g if x["is_night"]])
    row("Day Games (kick-off < 18:00)",   [x for x in g if x["is_day"]], alt=True)

    # ── DAY OF WEEK ──────────────────────────────────────────────
    section("DAY OF WEEK")
    row("Monday",                [x for x in g if x["is_monday"]])
    row("Tuesday / Wednesday",   [x for x in g if x["is_tue_wed"]], alt=True)
    row("Thursday",              [x for x in g if x["is_thursday"]])
    row("Friday",                [x for x in g if x["is_friday"]], alt=True)
    row("Saturday",              [x for x in g if x["is_saturday"]])
    row("Sunday",                [x for x in g if x["is_sunday"]], alt=True)

    # ── BY MONTH ─────────────────────────────────────────────────
    section("BY MONTH")
    epl_month_order = [8, 9, 10, 11, 12, 1, 2, 3, 4, 5]
    for i, m in enumerate(epl_month_order):
        row(MONTH_NAMES[m], [x for x in g if x["month"] == m], alt=(i % 2 == 1))

    # ── FORM ─────────────────────────────────────────────────────
    section("FORM")
    k_prev = f"prev_result__{team}"
    row("After a Win",  [x for x in g if x.get(k_prev) == "H"])
    row("After a Draw", [x for x in g if x.get(k_prev) == "D"], alt=True)
    row("After a Loss", [x for x in g if x.get(k_prev) == "A"])

    # ── REST ─────────────────────────────────────────────────────
    section("REST")
    k_rest = f"rest__{team}"
    row("Short Rest (≤ 6 days)",  [x for x in g if x.get(k_rest) is not None and x[k_rest] <= 6])
    row("Long Rest (≥ 10 days)",  [x for x in g if x.get(k_rest) is not None and x[k_rest] >= 10], alt=True)

    # ── MOON PHASE ───────────────────────────────────────────────
    section("MOON PHASE")
    row("New Moon (±1 day)",  [x for x in g if x["is_new_moon"]])
    row("Full Moon (±1 day)", [x for x in g if x["is_full_moon"]], alt=True)

    # ── HEAD TO HEAD ─────────────────────────────────────────────
    section("HEAD TO HEAD vs OPPONENT")
    opponents = sorted(t for t in all_teams if t != team)
    for i, opp in enumerate(opponents):
        opp_games = [x for x in g if x["home_team"] == opp or x["away_team"] == opp]
        row(f"vs {opp}", opp_games, alt=(i % 2 == 1))

    ws.freeze_panes = "B3"


# ─────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────

def main():
    print("Loading EPL data from CSVs...")
    all_rows = load_data()
    print(f"  Loaded {len(all_rows)} games")

    from collections import Counter
    for s, cnt in sorted(Counter(r["season"] for r in all_rows).items()):
        print(f"    {s}: {cnt} games")

    print("Enriching (moon, rest, form)...")
    all_rows = enrich_rows(all_rows)

    all_teams = sorted(set(r["home_team"] for r in all_rows) | set(r["away_team"] for r in all_rows))
    print(f"  {len(all_teams)} teams")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Building sheets...")
    for team in all_teams:
        g_count = sum(1 for r in all_rows if r["home_team"] == team or r["away_team"] == team)
        print(f"  {team} ({g_count} games)...")
        build_team_sheet(wb, team, all_rows, all_teams)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"Sheets ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}")

    # Summary of flagged edges
    flagged = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in ws.iter_rows(min_row=3):
            ec = r[4]
            if ec.value and isinstance(ec.value, str) and "edge" in ec.value:
                try:
                    if float(ec.value.split("%")[0]) >= EDGE_FLAG_PCT:
                        flagged += 1
                except Exception:
                    pass
    print(f"Flagged edges (≥{EDGE_FLAG_PCT}%): {flagged}")


if __name__ == "__main__":
    main()
